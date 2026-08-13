from flask import (
    Flask,
    request,
    render_template,
    redirect,
    flash,
    Response,
    send_file,
    session
)

from functools import wraps
from io import BytesIO

import os
import psycopg2

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


app = Flask(__name__)
app.secret_key = "edutrack_secret_key"

def get_db_connection():
    DATABASE_URL = os.getenv("DATABASE_URL")
    return psycopg2.connect(DATABASE_URL)


def calculate_average(marks):
    valid_marks = [mark for mark in marks if mark is not None]

    if not valid_marks:
        return 0

    return round(sum(valid_marks) / len(valid_marks), 2)


def calculate_grade(average):

    if average >= 90:
        return "A1"
    elif average >= 80:
        return "A2"
    elif average >= 70:
        return "B1"
    elif average >= 60:
        return "B2"
    elif average >= 50:
        return "C1"
    elif average >= 40:
        return "C2"
    elif average >= 33:
        return "D"
    else:
        return "F"
        
def admin_required(f):

    @wraps(f)
    def decorated_function(*args, **kwargs):

        if "admin" not in session:

            flash("Please login as Admin first.", "warning")

            return redirect("/login")

        return f(*args, **kwargs)

    return decorated_function

@app.route("/")
def index():

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            roll_number,
            name,
            math,
            science,
            english
        FROM students
    """)

    students = cur.fetchall()

    cur.close()
    conn.close()

    student_data = []

    for student in students:

        average = calculate_average([
            student[2],
            student[3],
            student[4]
        ])

        grade = calculate_grade(average)

        student_data.append({
            "roll": student[0],
            "name": student[1],
            "math": student[2],
            "science": student[3],
            "english": student[4],
            "average": average,
            "grade": grade
        })

# Highest average first
    student_data.sort(
        key=lambda x: x["average"],
        reverse=True
    )

    total_students = len(student_data)

    top_students = student_data[:10]

    averages = [
        student["average"]
        for student in student_data
        if student["average"] > 0
    ]

    class_average = (
        round(sum(averages) / len(averages), 2)
        if averages
        else 0
    )

    highest = (
        max(averages)
        if averages
        else 0
    )

    subjects = 3

    return render_template(
        "students/list.html",
        students=top_students,
        total_students=total_students,
        class_average=class_average,
        highest=highest,
        subjects=subjects
    )
    
@app.route("/view_students")
def view_students():

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            roll_number,
            name,
            math,
            science,
            english
        FROM students
        ORDER BY roll_number
    """)

    students = cur.fetchall()

    cur.close()
    conn.close()

    student_data = []

    for student in students:

        average = calculate_average([
            student[2],
            student[3],
            student[4]
        ])

        grade = calculate_grade(average)

        student_data.append({
            "roll": student[0],
            "name": student[1],
            "math": student[2],
            "science": student[3],
            "english": student[4],
            "average": average,
            "grade": grade
        })

    return render_template(
        "students/view_students.html",
        students=student_data
    )

@app.route("/add_student", methods=["GET", "POST"])
@admin_required
def add_student():

    if request.method == "POST":

        conn = get_db_connection()
        cur = conn.cursor()

        try:

            cur.execute(
                """
                INSERT INTO students
                (roll_number, name, math, science, english)
                VALUES (%s, %s, %s, %s, %s)
                """,
                (
                    request.form["roll"],
                    request.form["name"],
                    request.form["math"],
                    request.form["science"],
                    request.form["english"]
                )
            )

            conn.commit()

            flash(
                "🎉 Student added successfully!",
                "success"
            )

            return redirect("/")

        except psycopg2.IntegrityError:

            conn.rollback()

            flash(
                "❌ Roll Number already exists!",
                "danger"
            )

            return redirect("/add_student")

        finally:

            cur.close()
            conn.close()

    return render_template("students/add.html")

@app.route("/student/<int:roll>")
def student_details(roll):

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            roll_number,
            name,
            math,
            science,
            english
        FROM students
        WHERE roll_number = %s
    """, (str(roll),))

    student = cur.fetchone()

    cur.close()
    conn.close()

    if not student:
        return "Student Not Found"

    average = calculate_average([
        student[2],
        student[3],
        student[4]
    ])

    grade = calculate_grade(average)

    student_data = {
        "roll": student[0],
        "name": student[1],
        "math": student[2],
        "science": student[3],
        "english": student[4],
        "average": average,
        "grade": grade
    }

    return render_template(
        "students/details.html",
        student=student_data
    )
    
@app.route("/edit_student/<roll>", methods=["GET", "POST"])
@admin_required
def edit_student(roll):

    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == "POST":

        cur.execute(
            """
            UPDATE students
            SET
                name=%s,
                math=%s,
                science=%s,
                english=%s
            WHERE roll_number=%s
            """,
            (
                request.form["name"],
                request.form["math"],
                request.form["science"],
                request.form["english"],
                roll
            )
        )

        conn.commit()

        flash(
            "✅ Student updated successfully!",
            "success"
        )

        cur.close()
        conn.close()

        return redirect("/")

    cur.execute(
        """
        SELECT
            roll_number,
            name,
            math,
            science,
            english
        FROM students
        WHERE roll_number=%s
        """,
        (str(roll),)
    )

    student = cur.fetchone()

    cur.close()
    conn.close()

    if not student:
        return "Student Not Found"

    student_data = {
        "roll": student[0],
        "name": student[1],
        "math": student[2],
        "science": student[3],
        "english": student[4]
    }

    return render_template(
        "students/edit.html",
        student=student_data
    )

@app.route("/delete_student/<roll>", methods=["GET", "POST"])
@admin_required
def delete_student(roll):

    conn = get_db_connection()
    cur = conn.cursor()

    # Find the student first
    cur.execute("""
        SELECT
            roll_number,
            name,
            math,
            science,
            english
        FROM students
        WHERE roll_number = %s
    """, (str(roll),))

    student = cur.fetchone()

    if not student:

        cur.close()
        conn.close()

        return "Student Not Found"

    student_data = {
        "roll": student[0],
        "name": student[1],
        "math": student[2],
        "science": student[3],
        "english": student[4]
    }

    # Delete only after confirmation
    if request.method == "POST":

        cur.execute(
            "DELETE FROM students WHERE roll_number = %s",
            (str(roll),)
        )

        conn.commit()

        cur.close()
        conn.close()

        flash(
            "🗑 Student deleted successfully!",
            "success"
        )

        return redirect("/")

    cur.close()
    conn.close()

    return render_template(
        "students/delete.html",
        student=student_data
    )
@app.route("/export_csv")
@admin_required
def export_csv():

    DATABASE_URL = os.getenv("DATABASE_URL")

    conn = psycopg2.connect(DATABASE_URL)

    cur = conn.cursor()

    cur.execute("""
        SELECT
            roll_number,
            name,
            math,
            science,
            english
        FROM students
        ORDER BY roll_number
    """)

    students = cur.fetchall()

    conn.close()

    def generate():

        yield "Roll Number,Name,Math,Science,English\n"

        for s in students:

            yield f"{s[0]},{s[1]},{s[2]},{s[3]},{s[4]}\n"

    return Response(
        generate(),
        mimetype="text/csv",
        headers={
            "Content-Disposition":
            "attachment; filename=students.csv"
        }
    )

@app.route("/export_excel")
@admin_required
def export_excel():

    DATABASE_URL = os.getenv("DATABASE_URL")

    conn = psycopg2.connect(DATABASE_URL)
    cur = conn.cursor()

    cur.execute("""
        SELECT
            roll_number,
            name,
            math,
            science,
            english
        FROM students
        ORDER BY roll_number
    """)

    students = cur.fetchall()

    conn.close()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Students"

    sheet.append([
        "Roll Number",
        "Name",
        "Math",
        "Science",
        "English",
        "Average",
        "Grade"
    ])

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="0D6EFD"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    center = Alignment(
        horizontal="center",
        vertical="center"
    )

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for student in students:
        marks = [
            m for m in [
                student[2],
                student[3],
                student[4]
            ]
            if m is not None
        ]

        average = round(sum(marks) / len(marks), 2) if marks else 0

        if average >= 90:
            grade = "A1"
        elif average >= 80:
            grade = "A2"
        elif average >= 70:
            grade = "B1"
        elif average >= 60:
            grade = "B2"
        elif average >= 50:
            grade = "C1"
        elif average >= 40:
            grade = "C2"
        elif average >= 33:
            grade = "D"
        else:
            grade = "F"

        sheet.append([
            student[0],
            student[1],
            student[2],
            student[3],
            student[4],
            average,
            grade
        ])

    for column_cells in sheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        sheet.column_dimensions[column_letter].width = max_length + 5

    sheet.freeze_panes = "A2"

    sheet.auto_filter.ref = sheet.dimensions

    left = Alignment(horizontal="left")
    center = Alignment(horizontal="center")

    for row in sheet.iter_rows(min_row=2):
        row[0].alignment = center
        row[1].alignment = left
        row[2].alignment = center
        row[3].alignment = center
        row[4].alignment = center
        row[5].alignment = center
        row[6].alignment = center

    excel_file = BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)

    return send_file(
        excel_file,
        download_name="students.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument/spreadsheetml.sheet"
    )

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]

        if username == "admin" and password == "edutrack123":
            session["admin"] = username
            flash("Login Successful!", "success")
            return redirect("/")
        else:
            flash("Invalid Username or Password!", "danger")

    return render_template("auth/login.html")

@app.route("/logout")
def logout():

    session.pop("admin", None)

    flash("Logged out successfully!", "success")

    return redirect("/")

if __name__ == "__main__":
    app.run(debug=True)
    
